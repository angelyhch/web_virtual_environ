from openpyxl import Workbook
wb = Workbook()
# 激活 worksheet
ws = wb.active
# 数据可以直接分配到单元格中
ws['A1'] = 42
# 可以附加行，从第一列开始附加
ws.append([1, 2, 3])
# Python 类型会被自动转换
import datetime
ws['A3'] = datetime.datetime.now().strftime("%Y-%m-%d")
# 保存文件
wb.save("sample.xlsx")

